from pathlib import Path

import pytest
from openpyxl import Workbook

from afterlife_ai.contracts import (
    ProductCategory,
    StorageType,
    UnitCode,
    VerificationStatus,
)
from afterlife_ai.contracts.spreadsheet import REQUIRED_COLUMNS, SHEET_NAME
from afterlife_ai.intake.xlsx_reader import read_inventory_workbook


def valid_row() -> dict[str, object]:
    return {
        "lot_id": "LOT-001",
        "sku": "SKU-001",
        "product_name": "Reference Product",
        "product_category": next(iter(ProductCategory)).value,
        "current_quantity": 10,
        "unit": next(iter(UnitCode)).value,
        "unit_cost": 1000,
        "normal_selling_price": 1500,
        "source_location": "STORE-01",
        "storage_type": next(iter(StorageType)).value,
        "verification_status": next(iter(VerificationStatus)).value,
    }


def save_workbook(
    path: Path,
    *,
    sheet_name: str = SHEET_NAME,
    headers: tuple[str, ...] = REQUIRED_COLUMNS,
    include_data: bool = True,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(list(headers))

    if include_data:
        row = valid_row()
        worksheet.append([row.get(header) for header in headers])

    workbook.save(path)
    workbook.close()


def test_missing_required_worksheet_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "missing_sheet.xlsx"
    save_workbook(path, sheet_name="wrong_sheet")

    with pytest.raises(ValueError, match="Worksheet wajib"):
        read_inventory_workbook(path)


def test_missing_required_column_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "missing_column.xlsx"
    headers = tuple(
        column
        for column in REQUIRED_COLUMNS
        if column != "verification_status"
    )
    save_workbook(path, headers=headers)

    with pytest.raises(ValueError, match="Kolom wajib tidak ditemukan"):
        read_inventory_workbook(path)


def test_workbook_without_inventory_rows_is_rejected(
    tmp_path: Path,
) -> None:
    path = tmp_path / "empty_inventory.xlsx"
    save_workbook(path, include_data=False)

    with pytest.raises(
        ValueError,
        match="tidak memiliki baris inventori",
    ):
        read_inventory_workbook(path)


def test_unknown_column_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "unknown_column.xlsx"
    headers = (*REQUIRED_COLUMNS, "invented_column")
    save_workbook(path, headers=headers)

    with pytest.raises(ValueError, match="Kolom tidak dikenal"):
        read_inventory_workbook(path)


def test_extension_column_is_accepted(tmp_path: Path) -> None:
    path = tmp_path / "extension_column.xlsx"
    headers = (*REQUIRED_COLUMNS, "x_operator_note")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(list(headers))

    row = valid_row()
    worksheet.append(
        [
            *[row[column] for column in REQUIRED_COLUMNS],
            "manual review note",
        ]
    )

    workbook.save(path)
    workbook.close()

    records = read_inventory_workbook(path)

    assert len(records) == 1
    assert records[0].lot_id == "LOT-001"


def test_exact_duplicate_rows_are_rejected(tmp_path: Path) -> None:
    path = tmp_path / "duplicate_rows.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(list(REQUIRED_COLUMNS))

    row = valid_row()
    values = [row[column] for column in REQUIRED_COLUMNS]

    worksheet.append(values)
    worksheet.append(values)

    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="Baris duplikat"):
        read_inventory_workbook(path)
