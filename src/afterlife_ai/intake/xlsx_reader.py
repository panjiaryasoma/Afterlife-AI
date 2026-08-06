"""Read inventory XLSX workbooks into validated inventory contracts."""

from pathlib import Path

from openpyxl import load_workbook

from afterlife_ai.contracts.inventory import RawInventoryLot
from afterlife_ai.contracts.spreadsheet import (
    ALL_CANONICAL_COLUMNS,
    EXTENSION_COLUMN_PREFIX,
    REQUIRED_COLUMNS,
    SHEET_NAME,
)


def read_inventory_workbook(file_path: str | Path) -> list[RawInventoryLot]:
    """Read one inventory workbook and return validated inventory lots."""

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Workbook tidak ditemukan: {path}")

    if not path.is_file():
        raise ValueError(f"Path bukan file: {path}")

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet wajib '{SHEET_NAME}' tidak ditemukan. "
                f"Worksheet tersedia: {workbook.sheetnames}"
            )

        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)

        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Workbook tidak memiliki header atau data.") from exc

        headers = tuple(
            value.strip() if isinstance(value, str) else value
            for value in raw_headers
        )

        if any(header is None or header == "" for header in headers):
            raise ValueError("Header workbook tidak boleh kosong.")

        if not all(isinstance(header, str) for header in headers):
            raise ValueError("Seluruh nama kolom harus berupa teks.")

        duplicate_headers = {
            header for header in headers if headers.count(header) > 1
        }
        if duplicate_headers:
            duplicates = ", ".join(sorted(duplicate_headers))
            raise ValueError(f"Nama kolom duplikat ditemukan: {duplicates}")

        missing_columns = [
            column for column in REQUIRED_COLUMNS if column not in headers
        ]
        if missing_columns:
            missing = ", ".join(missing_columns)
            raise ValueError(f"Kolom wajib tidak ditemukan: {missing}")

        unknown_columns = [
            header
            for header in headers
            if header not in ALL_CANONICAL_COLUMNS
            and not header.startswith(EXTENSION_COLUMN_PREFIX)
        ]
        if unknown_columns:
            unknown = ", ".join(sorted(unknown_columns))
            raise ValueError(f"Kolom tidak dikenal: {unknown}")

        records: list[RawInventoryLot] = []

        for row_number, row in enumerate(rows, start=2):
            if all(
                value is None
                or (isinstance(value, str) and not value.strip())
                for value in row
            ):
                continue

            payload = {
                header: value
                for header, value in zip(headers, row, strict=True)
                if header in ALL_CANONICAL_COLUMNS
            }

            try:
                record = RawInventoryLot.model_validate(payload)
            except Exception as exc:
                raise ValueError(
                    f"Baris {row_number} tidak valid: {exc}"
                ) from exc

            records.append(record)

        if not records:
            raise ValueError("Workbook tidak memiliki baris inventori yang valid.")

        return records

    finally:
        workbook.close()


__all__ = ["read_inventory_workbook"]
